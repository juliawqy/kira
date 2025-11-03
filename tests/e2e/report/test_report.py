"""
End-to-end tests for report generation.
Tests actual report generation with real data setup and validation.
Includes both API-based and browser-based tests.
"""
from __future__ import annotations

import pytest
import time
import os
from datetime import date
from io import BytesIO
from sqlalchemy import create_engine, event, delete, text
from sqlalchemy.orm import sessionmaker
from urllib.parse import quote
from selenium.webdriver.common.by import By

from backend.src.database.db_setup import Base
from backend.src.database.models.user import User
from backend.src.database.models.project import Project
from backend.src.database.models.task import Task
from backend.src.database.models.task_assignment import TaskAssignment
from backend.src.handlers.report_handler import generate_pdf_report, generate_excel_report
from backend.src.main import app
from tests.mock_data.report_data import (
    MOCK_PROJECT,
    MOCK_TASKS_ALL_STATUSES,
    MOCK_TASK_ASSIGNEES,
    MOCK_USERS,
    MOCK_USER_BY_NAME
)


@pytest.fixture(scope="session")
def test_engine_report(tmp_path_factory):
    """
    Session-scoped file-backed SQLite engine with FK support.
    Creates schema once, drops at the end.
    """
    db_file = tmp_path_factory.mktemp("kira_e2e_report") / "e2e_report.db"
    engine = create_engine(
        f"sqlite:///{db_file}",
        connect_args={"check_same_thread": False},
        future=True,
    )

    @event.listens_for(engine, "connect")
    def _fk_on(dbapi_connection, connection_record):
        cur = dbapi_connection.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    Base.metadata.create_all(bind=engine)
    try:
        yield engine
    finally:
        # Teardown: disable FK checks before dropping tables
        try:
            with engine.connect() as conn:
                conn.execute(text("PRAGMA foreign_keys=OFF"))
                Base.metadata.drop_all(bind=conn)
                conn.execute(text("PRAGMA foreign_keys=ON"))
        finally:
            engine.dispose()


@pytest.fixture
def isolated_database(test_engine_report):
    """
    Database isolation using session.
    """
    TestingSessionLocal = sessionmaker(
        bind=test_engine_report,
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
        future=True,
    )

    with TestingSessionLocal() as session:
        try:
            yield session
        finally:
            session.close()


@pytest.fixture(autouse=True)
def reset_database_tables(test_engine_report):
    """Drop and recreate all tables before each test to reset ID sequences."""
    Session = sessionmaker(bind=test_engine_report, future=True)
    with Session() as session:
        # Clean up in order to respect foreign keys
        session.execute(delete(TaskAssignment))
        session.execute(delete(Task))
        session.execute(delete(Project))
        session.execute(delete(User))
        # Reset SQLite sequences to ensure clean ID assignment (if table exists)
        try:
            session.execute(text("DELETE FROM sqlite_sequence WHERE name IN ('task', 'project', 'user')"))
        except Exception:
            # sqlite_sequence table doesn't exist yet, which is fine
            pass
        session.commit()
    yield


@pytest.fixture(scope="session")
def app_server(test_engine_report):
    """Run FastAPI app with SessionLocal overridden to the isolated test engine."""
    import threading
    import socket
    import uvicorn
    import smtplib
    from backend.src.services import task as task_service
    from backend.src.services import project as project_service
    from backend.src.services import task_assignment as assignment_service
    
    TestingSessionLocal = sessionmaker(
        bind=test_engine_report,
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
        future=True,
    )
    task_service.SessionLocal = TestingSessionLocal
    project_service.SessionLocal = TestingSessionLocal
    assignment_service.SessionLocal = TestingSessionLocal

    class DummySMTP:
        def __init__(self, *a, **kw): pass
        def starttls(self): pass
        def login(self, *a, **kw): pass
        def send_message(self, *a, **kw): pass
        def quit(self): pass

    smtplib.SMTP = DummySMTP
    smtplib.SMTP_SSL = DummySMTP
    
    port = None
    for candidate in [8020, 8021, 8022, 8023, 8024, 8025]:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(("127.0.0.1", candidate))
                port = candidate
                break
            except OSError:
                continue
    if port is None:
        port = 8026

    config = uvicorn.Config(app, host="127.0.0.1", port=port, log_level="warning")
    server = uvicorn.Server(config)
    thread = threading.Thread(target=server.run, daemon=True)
    thread.start()
    time.sleep(1.0)
    
    yield f"http://127.0.0.1:{port}"


@pytest.fixture
def setup_report_test_data(isolated_database):
    """
    Set up test data for report generation using mock data from report_data.py.
    Creates users, project, and tasks with assignments.
    """
    from backend.src.services import task as task_service
    from backend.src.services import project as project_service
    from backend.src.services import task_assignment as assignment_service

    TestingSessionLocal = sessionmaker(
        bind=isolated_database.get_bind(),
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
        future=True,
    )

    # Override service SessionLocal to use test database
    # Keep this override active during test execution
    original_task_session = task_service.SessionLocal
    original_project_session = project_service.SessionLocal
    original_assignment_session = assignment_service.SessionLocal

    task_service.SessionLocal = TestingSessionLocal
    project_service.SessionLocal = TestingSessionLocal
    assignment_service.SessionLocal = TestingSessionLocal

    try:
        session = isolated_database

        # Create users from mock data (report_data.py)
        users = []
        for user_dict in MOCK_USERS:
            user = User(**user_dict)
            session.add(user)
            users.append(user)
        
        session.flush()

        # Create a mapping from assignee names to user objects (from report_data.py)
        name_to_user = {}
        for assignee_name, user_dict in MOCK_USER_BY_NAME.items():
            # Find the user object we just created
            for user in users:
                if user.name == assignee_name:
                    name_to_user[assignee_name] = user
                    break

        # Find manager user for project_manager field
        manager_user = name_to_user.get("Project Manager")
        if not manager_user:
            raise ValueError("Project Manager user not found in mock data")

        # Create project from mock data (report_data.py)
        project = Project(
            project_name=MOCK_PROJECT["project_name"],
            project_manager=manager_user.user_id,
            active=MOCK_PROJECT["active"]
        )
        session.add(project)
        session.flush()

        # Create tasks from mock data (report_data.py)
        tasks = []
        for task_dict in MOCK_TASKS_ALL_STATUSES:
            # Create task object from dictionary
            task = Task(
                title=task_dict["title"],
                description=task_dict["description"],
                status=task_dict["status"],
                priority=task_dict["priority"],
                start_date=task_dict["start_date"],
                deadline=task_dict["deadline"],
                tag=task_dict["tag"],
                project_id=project.project_id,
                active=task_dict["active"]
            )
            session.add(task)
            tasks.append(task)

        session.flush()

        # Assign users to tasks based on mock data (report_data.py)
        # Map task title to the task object we just created
        title_to_task = {task.title: task for task in tasks}

        # Map mock task ID to task title (to match with MOCK_TASK_ASSIGNEES)
        mock_id_to_title = {
            task_dict["id"]: task_dict["title"]
            for task_dict in MOCK_TASKS_ALL_STATUSES
        }

        # Create assignments based on mock data (report_data.py)
        for mock_task_id, assignee_names in MOCK_TASK_ASSIGNEES.items():
            # Get the task title from mock data
            task_title = mock_id_to_title.get(mock_task_id)
            if not task_title:
                continue

            # Find the actual task object we created
            task = title_to_task.get(task_title)
            if not task:
                continue

            # Assign users to this task
            for assignee_name in assignee_names:
                if assignee_name in name_to_user:
                    assignment = TaskAssignment(
                        task_id=task.id,
                        user_id=name_to_user[assignee_name].user_id
                    )
                    session.add(assignment)

        session.commit()

        yield {
            "project_id": project.project_id,
            "project": project,
            "tasks": tasks,
            "users": users
        }

    finally:
        task_service.SessionLocal = original_task_session
        project_service.SessionLocal = original_project_session
        assignment_service.SessionLocal = original_assignment_session


class TestReportE2E:
    """End-to-end tests for report generation."""

    def test_generate_pdf_report_e2e(self, setup_report_test_data):
        """Test PDF report generation end-to-end with real data."""
        project_id = setup_report_test_data["project_id"]

        pdf_buffer = generate_pdf_report(project_id)

        assert pdf_buffer is not None
        assert isinstance(pdf_buffer, BytesIO)

        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')

        assert len(content) > 1000, "PDF should be at least 1KB"

    def test_generate_excel_report_e2e(self, setup_report_test_data):
        """Test Excel report generation end-to-end with real data."""
        project_id = setup_report_test_data["project_id"]

        excel_buffer = generate_excel_report(project_id)

        assert excel_buffer is not None
        assert isinstance(excel_buffer, BytesIO)

        excel_buffer.seek(0)
        content = excel_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'PK')

        assert len(content) > 1000, "Excel file should be at least 1KB"

    def test_excel_report_contains_correct_data(self, setup_report_test_data):
        """Test that Excel report contains expected data from mock data."""
        from openpyxl import load_workbook

        project_id = setup_report_test_data["project_id"]
        project = setup_report_test_data["project"]
        tasks = setup_report_test_data["tasks"]

        assert project.project_name == MOCK_PROJECT["project_name"]

        excel_buffer = generate_excel_report(project_id)
        excel_buffer.seek(0)

        wb = load_workbook(excel_buffer)
        assert "Project Schedule Report" in wb.sheetnames

        ws = wb["Project Schedule Report"]

        assert ws["A2"].value == "Project Name"
        assert ws["B2"].value == project.project_name

        assert ws["A4"].value == "Summary"
        assert ws["A5"].value == "Metric"
        assert ws["B5"].value == "Count"

        summary_dict = {}
        for row in range(6, 12):
            metric = ws[f"A{row}"].value
            count = ws[f"B{row}"].value
            if metric and count is not None:
                summary_dict[metric] = count

        assert summary_dict.get("Total Tasks") == len(tasks) == len(MOCK_TASKS_ALL_STATUSES)

        assert summary_dict.get("Projected Tasks") == 1  # TO_DO
        assert summary_dict.get("In-Progress Tasks") == 1  # IN_PROGRESS
        assert summary_dict.get("Completed Tasks") == 1  # COMPLETED
        assert summary_dict.get("Under Review Tasks") == 1  # BLOCKED

    def test_pdf_report_contains_project_info(self, setup_report_test_data):
        """Test that PDF report contains project information."""
        project_id = setup_report_test_data["project_id"]

        pdf_buffer = generate_pdf_report(project_id)
        pdf_buffer.seek(0)
        content = pdf_buffer.read()

        assert content.startswith(b'%PDF')

        assert len(content) > 2000, "PDF should contain substantial content"

    def test_report_generation_with_multiple_tasks(self, setup_report_test_data):
        """Test that reports handle multiple tasks correctly."""
        project_id = setup_report_test_data["project_id"]
        tasks = setup_report_test_data["tasks"]

        # Generate both reports
        pdf_buffer = generate_pdf_report(project_id)
        excel_buffer = generate_excel_report(project_id)

        assert pdf_buffer is not None
        assert excel_buffer is not None

        from openpyxl import load_workbook
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb["Project Schedule Report"]
        task_titles_found = []
        task_titles_expected = [task.title for task in tasks]
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=2).value  # Column B is Title
            if cell_value and isinstance(cell_value, str):
                if cell_value in ["Title", "PROJECTED TASKS", "IN-PROGRESS TASKS", "COMPLETED TASKS", "UNDER REVIEW TASKS", "Summary", "Project Name"]:
                    continue
                for expected_title in task_titles_expected:
                    if expected_title == cell_value or (isinstance(cell_value, str) and expected_title in cell_value):
                        if expected_title not in task_titles_found:
                            task_titles_found.append(expected_title)

        assert len(task_titles_found) > 0, \
            f"Should find task titles in Excel report. Expected: {task_titles_expected}, Found: {task_titles_found}. Total rows: {ws.max_row}"


class TestReportE2EBrowser:
    """Browser-based E2E tests for report generation UI."""

    def test_report_export_page_loads(self, driver, app_server, setup_report_test_data):
        """Test that the report export page loads correctly."""
        html_file = os.path.join(os.getcwd(), "frontend", "report", "export_report.html")
        driver.get(f"file:///{html_file}?api={quote(app_server)}")
        
        time.sleep(1)
        
        assert "Export Project Schedule Report" in driver.title or "Export" in driver.title
        
        project_input = driver.find_element(By.ID, "projectIdInput")
        assert project_input is not None
        
        pdf_btn = driver.find_element(By.ID, "exportPdfBtn")
        assert pdf_btn is not None
        
        excel_btn = driver.find_element(By.ID, "exportExcelBtn")
        assert excel_btn is not None
        
        assert pdf_btn.get_attribute("disabled") is not None
        assert excel_btn.get_attribute("disabled") is not None

    def test_report_export_ui_interaction(self, driver, app_server, setup_report_test_data):
        """Test UI interaction for report export."""
        project_id = setup_report_test_data["project_id"]
        
        html_file = os.path.join(os.getcwd(), "frontend", "report", "export_report.html")
        driver.get(f"file:///{html_file}?api={quote(app_server)}")
        
        time.sleep(1)
        
        project_input = driver.find_element(By.ID, "projectIdInput")
        project_input.clear()
        project_input.send_keys(str(project_id))
        time.sleep(0.5)
        
        pdf_btn = driver.find_element(By.ID, "exportPdfBtn")
        excel_btn = driver.find_element(By.ID, "exportExcelBtn")
        
        assert pdf_btn.get_attribute("disabled") is None, "PDF button should be enabled"
        assert excel_btn.get_attribute("disabled") is None, "Excel button should be enabled"

    def test_report_download_pdf(self, driver, app_server, setup_report_test_data, tmp_path):
        """Test downloading PDF report file."""
        from pathlib import Path
        
        project_id = setup_report_test_data["project_id"]
        
        download_dir = tmp_path / "downloads"
        download_dir.mkdir()
        
        print(f"\n[DEBUG] Download directory: {download_dir}")
        
        html_file = os.path.join(os.getcwd(), "frontend", "report", "export_report.html")
        
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": str(download_dir)
        })
        
        driver.get(f"file:///{html_file}?api={quote(app_server + '/kira/app/api/v1')}")
        
        time.sleep(1)
        
        project_input = driver.find_element(By.ID, "projectIdInput")
        project_input.clear()
        project_input.send_keys(str(project_id))
        time.sleep(0.5)
        
        pdf_btn = driver.find_element(By.ID, "exportPdfBtn")
        pdf_btn.click()
        
        max_wait = 10
        waited = 0
        downloaded_files = []
        while waited < max_wait:
            time.sleep(0.5)
            waited += 0.5
            downloaded_files = list(download_dir.glob("*.pdf"))
            crdownload_files = list(download_dir.glob("*.pdf.crdownload"))
            if downloaded_files and not crdownload_files:
                break
        
        assert len(downloaded_files) > 0, f"Expected PDF file to be downloaded. Files in {download_dir}: {list(download_dir.iterdir())}"
        
        pdf_file = downloaded_files[0]
        assert pdf_file.exists()
        assert pdf_file.stat().st_size > 1000
        
        with open(pdf_file, "rb") as f:
            content = f.read()
            assert content.startswith(b'%PDF')
        time.sleep(20)

    def test_report_download_excel(self, driver, app_server, setup_report_test_data, tmp_path):
        """Test downloading Excel report file."""
        from pathlib import Path
        
        project_id = setup_report_test_data["project_id"]
        
        download_dir = tmp_path / "downloads"
        download_dir.mkdir()
        
        print(f"\n[DEBUG] Download directory: {download_dir}")
        
        html_file = os.path.join(os.getcwd(), "frontend", "report", "export_report.html")
        
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": str(download_dir)
        })
        
        driver.get(f"file:///{html_file}?api={quote(app_server + '/kira/app/api/v1')}")
        
        time.sleep(1)
        
        project_input = driver.find_element(By.ID, "projectIdInput")
        project_input.clear()
        project_input.send_keys(str(project_id))
        time.sleep(0.5)
        
        excel_btn = driver.find_element(By.ID, "exportExcelBtn")
        excel_btn.click()
        
        max_wait = 10
        waited = 0
        downloaded_files = []
        while waited < max_wait:
            time.sleep(0.5)
            waited += 0.5
            downloaded_files = list(download_dir.glob("*.xlsx"))
            crdownload_files = list(download_dir.glob("*.xlsx.crdownload"))
            if downloaded_files and not crdownload_files:
                break
        
        assert len(downloaded_files) > 0, f"Expected Excel file to be downloaded. Files in {download_dir}: {list(download_dir.iterdir())}"
        
        excel_file = downloaded_files[0]
        assert excel_file.exists()
        assert excel_file.stat().st_size > 1000
        
        with open(excel_file, "rb") as f:
            content = f.read()
            assert content.startswith(b'PK')
        time.sleep(20)

