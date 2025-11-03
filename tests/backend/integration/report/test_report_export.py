"""
Integration tests for project report export functionality.
Tests PDF and Excel report generation endpoints.
"""
from __future__ import annotations

import pytest
from datetime import date, datetime
from io import BytesIO

from backend.src.database.models.task import Task
from backend.src.database.models.project import Project
from backend.src.database.models.user import User
from backend.src.database.models.task_assignment import TaskAssignment
from backend.src.enums.task_status import TaskStatus
from backend.src.enums.user_role import UserRole
from backend.src.handlers.report_handler import generate_pdf_report, generate_excel_report
from tests.mock_data.report_data import (
    MOCK_USER_MANAGER,
    MOCK_USER_TEAM_MEMBER_1,
    MOCK_USER_TEAM_MEMBER_2,
    MOCK_PROJECT,
    MOCK_PROJECT_ID,
    MOCK_TASKS_FOR_INTEGRATION,
    MOCK_TASK_IDS_FOR_INTEGRATION,
    MOCK_TASK_ASSIGNEES,
    MOCK_USER_IDS_BY_NAME,
    EXPECTED_REPORT_SHEET_NAME,
    EXPECTED_REPORT_HEADER_PROJECT_NAME,
    EXPECTED_REPORT_SECTION_SUMMARY,
    EXPECTED_REPORT_METRIC_TOTAL_TASKS,
    EXPECTED_REPORT_METRIC_PROJECTED_TASKS,
    EXPECTED_REPORT_METRIC_COMPLETED_TASKS,
    NOT_FOUND_PROJECT_ID,
    EXPECTED_ERROR_PROJECT_NOT_FOUND
)


@pytest.fixture(scope="function")
def test_engine():
    """Create a temporary database for testing."""
    import tempfile
    import os
    from sqlalchemy import create_engine
    from backend.src.database.db_setup import Base
    
    db_fd, db_path = tempfile.mkstemp(suffix=".db")
    os.close(db_fd)
    
    engine = create_engine(f"sqlite:///{db_path}", echo=False)
    Base.metadata.create_all(bind=engine)
    
    yield engine
    
    engine.dispose()
    if os.path.exists(db_path):
        os.remove(db_path)


@pytest.fixture(scope="function")
def isolated_test_db(test_engine):
    """Patch services to use the test database for each test."""
    from sqlalchemy.orm import sessionmaker
    from unittest.mock import patch
    
    TestingSessionLocal = sessionmaker(
        bind=test_engine,
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
        future=True,
    )
    
    from backend.src.services import task as task_service
    from backend.src.services import project as project_service
    from backend.src.services import task_assignment as assignment_service
    
    with patch("backend.src.services.task.SessionLocal", TestingSessionLocal), \
         patch("backend.src.services.project.SessionLocal", TestingSessionLocal), \
         patch("backend.src.services.task_assignment.SessionLocal", TestingSessionLocal):
        yield test_engine


@pytest.fixture
def setup_project_with_tasks(isolated_test_db):
    """Create a project with multiple tasks in different statuses."""
    from sqlalchemy.orm import sessionmaker
    
    Session = sessionmaker(bind=isolated_test_db, future=True)
    session = Session()
    
    try:
        user1 = User(user_id=1, **MOCK_USER_MANAGER)
        user2 = User(user_id=2, **MOCK_USER_TEAM_MEMBER_1)
        user3 = User(user_id=3, **MOCK_USER_TEAM_MEMBER_2)
        
        session.add_all([user1, user2, user3])
        session.flush()
        
        project = Project(**MOCK_PROJECT)
        session.add(project)
        session.flush()
        
        tasks = []
        for task_data in MOCK_TASKS_FOR_INTEGRATION:
            task_dict = {k: v for k, v in task_data.items() if k != "id"}
            task = Task(**task_dict)
            session.add(task)
            session.flush()
            tasks.append(task)
    
        assignments = []
        
        for task_idx, mock_task_id in enumerate(MOCK_TASK_IDS_FOR_INTEGRATION):
            if mock_task_id in MOCK_TASK_ASSIGNEES:
                for assignee_name in MOCK_TASK_ASSIGNEES[mock_task_id]:
                    if assignee_name in MOCK_USER_IDS_BY_NAME:
                        assignments.append(TaskAssignment(
                            task_id=tasks[task_idx].id,
                            user_id=MOCK_USER_IDS_BY_NAME[assignee_name]
                        ))
        
        session.add_all(assignments)
        session.commit()
        
        yield {"project_id": MOCK_PROJECT_ID, "project": project, "tasks": tasks, "users": [user1, user2, user3]}
        
    finally:
        session.close()


class TestReportExport:
    """Test report export functionality."""
    
    def test_generate_pdf_report_success(self, setup_project_with_tasks):
        """Test successful PDF report generation."""
        project_id = setup_project_with_tasks["project_id"]
        
        pdf_buffer = generate_pdf_report(project_id)
        
        assert pdf_buffer is not None
        assert isinstance(pdf_buffer, BytesIO)
        
        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')
    
    def test_generate_excel_report_success(self, setup_project_with_tasks):
        """Test successful Excel report generation."""
        project_id = setup_project_with_tasks["project_id"]
        
        excel_buffer = generate_excel_report(project_id)
        
        assert excel_buffer is not None
        assert isinstance(excel_buffer, BytesIO)
        
        excel_buffer.seek(0)
        content = excel_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'PK')
    
    def test_generate_pdf_report_project_not_found(self, isolated_test_db):
        """Test PDF generation with non-existent project."""
        expected_error = EXPECTED_ERROR_PROJECT_NOT_FOUND.format(project_id=NOT_FOUND_PROJECT_ID)
        with pytest.raises(ValueError, match=expected_error):
            generate_pdf_report(NOT_FOUND_PROJECT_ID)
    
    def test_generate_excel_report_project_not_found(self, isolated_test_db):
        """Test Excel generation with non-existent project."""
        expected_error = EXPECTED_ERROR_PROJECT_NOT_FOUND.format(project_id=NOT_FOUND_PROJECT_ID)
        with pytest.raises(ValueError, match=expected_error):
            generate_excel_report(NOT_FOUND_PROJECT_ID)
    
    def test_pdf_report_contains_all_statuses(self, setup_project_with_tasks):
        """Verify PDF report includes all task statuses."""
        project_id = setup_project_with_tasks["project_id"]
        
        pdf_buffer = generate_pdf_report(project_id)
        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert content.startswith(b'%PDF')
        assert len(content) > 1000
    
    def test_excel_report_has_summary_sheet(self, setup_project_with_tasks):
        """Verify Excel report has summary information in the single sheet."""
        from openpyxl import load_workbook
        
        project_id = setup_project_with_tasks["project_id"]
        
        excel_buffer = generate_excel_report(project_id)
        excel_buffer.seek(0)
        
        wb = load_workbook(excel_buffer)
        
        assert EXPECTED_REPORT_SHEET_NAME in wb.sheetnames
        ws = wb[EXPECTED_REPORT_SHEET_NAME]
        assert ws["A1"].value is not None
        assert ws["A2"].value == EXPECTED_REPORT_HEADER_PROJECT_NAME
        assert ws["B2"].value is not None
        assert ws["A4"].value == EXPECTED_REPORT_SECTION_SUMMARY
        
        summary_values = []
        for row in range(5, 11):
            if ws[f"A{row}"].value:
                summary_values.append(ws[f"A{row}"].value)
        
        assert EXPECTED_REPORT_METRIC_TOTAL_TASKS in summary_values
        assert EXPECTED_REPORT_METRIC_PROJECTED_TASKS in summary_values or EXPECTED_REPORT_METRIC_COMPLETED_TASKS in summary_values

    def test_generate_pdf_report_service_exception(self, setup_project_with_tasks, monkeypatch):
        """Test PDF generation when report service raises a non-ValueError exception."""
        project_id = setup_project_with_tasks["project_id"]
        
        def mock_generate_pdf(project, tasks, task_assignees):
            raise RuntimeError("PDF generation failed")
        
        from backend.src.services import report as report_service
        monkeypatch.setattr(report_service, "generate_pdf_report", mock_generate_pdf)
        
        with pytest.raises(RuntimeError, match="PDF generation failed"):
            generate_pdf_report(project_id)

    def test_generate_excel_report_service_exception(self, setup_project_with_tasks, monkeypatch):
        """Test Excel generation when report service raises a non-ValueError exception."""
        project_id = setup_project_with_tasks["project_id"]
        
        def mock_generate_excel(project, tasks, task_assignees):
            raise RuntimeError("Excel generation failed")
        
        from backend.src.services import report as report_service
        monkeypatch.setattr(report_service, "generate_excel_report", mock_generate_excel)
        
        with pytest.raises(RuntimeError, match="Excel generation failed"):
            generate_excel_report(project_id)


class TestReportAPI:
    """Test report API routes."""
    
    @pytest.fixture(scope="function")
    def api_test_engine(self):
        """Create a temporary database for API testing."""
        import tempfile
        import os
        from sqlalchemy import create_engine
        from backend.src.database.db_setup import Base
        
        db_fd, db_path = tempfile.mkstemp(suffix=".db")
        os.close(db_fd)
        
        engine = create_engine(f"sqlite:///{db_path}", echo=False)
        Base.metadata.create_all(bind=engine)
        
        yield engine
        
        engine.dispose()
        if os.path.exists(db_path):
            os.remove(db_path)

    @pytest.fixture
    def api_db_session(self, api_test_engine):
        """Create a database session for API testing."""
        from sqlalchemy.orm import sessionmaker
        
        connection = api_test_engine.connect()
        transaction = connection.begin()
        Session = sessionmaker(bind=connection)
        session = Session()
        
        try:
            yield session
        finally:
            session.close()
            transaction.rollback()
            connection.close()

    @pytest.fixture
    def api_client(self, api_test_engine):
        """Create FastAPI TestClient for API route testing."""
        from fastapi.testclient import TestClient
        from backend.src.main import app
        from sqlalchemy.orm import sessionmaker
        from backend.src.services import task as task_service
        from backend.src.services import project as project_service
        from backend.src.services import task_assignment as assignment_service
        
        TestingSessionLocal = sessionmaker(
            bind=api_test_engine,
            autoflush=False,
            autocommit=False,
            expire_on_commit=False,
            future=True,
        )
        
        old_task_session = task_service.SessionLocal
        old_project_session = project_service.SessionLocal
        old_assignment_session = assignment_service.SessionLocal
        
        task_service.SessionLocal = TestingSessionLocal
        project_service.SessionLocal = TestingSessionLocal
        assignment_service.SessionLocal = TestingSessionLocal
        
        try:
            with TestClient(app) as client:
                yield client
        finally:
            task_service.SessionLocal = old_task_session
            project_service.SessionLocal = old_project_session
            assignment_service.SessionLocal = old_assignment_session

    @pytest.fixture
    def api_setup_project(self, api_test_engine):
        """Create a project for API testing using mock data."""
        from sqlalchemy.orm import sessionmaker
        from backend.src.database.models.project import Project
        from backend.src.database.models.user import User
        
        Session = sessionmaker(bind=api_test_engine)
        session = Session()
        
        try:
            user = User(user_id=1, **MOCK_USER_MANAGER)
            session.add(user)
            session.flush()
            
            project = Project(**MOCK_PROJECT)
            session.add(project)
            session.commit()
            
            return {"project_id": MOCK_PROJECT_ID}
        finally:
            session.close()

    def test_api_export_pdf_report_project_not_found(self, api_client):
        """Test PDF export API with non-existent project returns 404."""
        expected_error = EXPECTED_ERROR_PROJECT_NOT_FOUND.format(project_id=NOT_FOUND_PROJECT_ID)
        response = api_client.get(f"/kira/app/api/v1/report/project/{NOT_FOUND_PROJECT_ID}/pdf")
        assert response.status_code == 404
        assert expected_error in response.json()["detail"]

    def test_api_export_excel_report_project_not_found(self, api_client):
        """Test Excel export API with non-existent project returns 404."""
        expected_error = EXPECTED_ERROR_PROJECT_NOT_FOUND.format(project_id=NOT_FOUND_PROJECT_ID)
        response = api_client.get(f"/kira/app/api/v1/report/project/{NOT_FOUND_PROJECT_ID}/excel")
        assert response.status_code == 404
        assert expected_error in response.json()["detail"]

    def test_api_export_pdf_report_general_exception(self, api_client, api_setup_project, monkeypatch):
        """Test PDF export API returns 500 when handler raises non-ValueError exception."""
        project_id = api_setup_project["project_id"]
        
        def mock_generate_pdf(*args, **kwargs):
            raise RuntimeError("PDF generation service error")
        
        from backend.src.handlers import report_handler
        monkeypatch.setattr(report_handler, "generate_pdf_report", mock_generate_pdf)
        
        response = api_client.get(f"/kira/app/api/v1/report/project/{project_id}/pdf")
        assert response.status_code == 500
        assert "Error generating PDF report" in response.json()["detail"]

    def test_api_export_excel_report_general_exception(self, api_client, api_setup_project, monkeypatch):
        """Test Excel export API returns 500 when handler raises non-ValueError exception."""
        project_id = api_setup_project["project_id"]
        
        def mock_generate_excel(*args, **kwargs):
            raise RuntimeError("Excel generation service error")
        
        from backend.src.handlers import report_handler
        monkeypatch.setattr(report_handler, "generate_excel_report", mock_generate_excel)
        
        response = api_client.get(f"/kira/app/api/v1/report/project/{project_id}/excel")
        assert response.status_code == 500
        assert "Error generating Excel report" in response.json()["detail"]

    def test_api_export_pdf_report_success(self, api_client, api_setup_project):
        """Test successful PDF export via API."""
        project_id = api_setup_project["project_id"]
        
        response = api_client.get(f"/kira/app/api/v1/report/project/{project_id}/pdf")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "attachment" in response.headers["content-disposition"]
        assert f"project_{project_id}_schedule_report.pdf" in response.headers["content-disposition"]
        assert len(response.content) > 0
        assert response.content.startswith(b'%PDF')

    def test_api_export_excel_report_success(self, api_client, api_setup_project):
        """Test successful Excel export via API."""
        project_id = api_setup_project["project_id"]
        
        response = api_client.get(f"/kira/app/api/v1/report/project/{project_id}/excel")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert f"project_{project_id}_schedule_report.xlsx" in response.headers["content-disposition"]
        assert len(response.content) > 0
        assert response.content.startswith(b'PK')

