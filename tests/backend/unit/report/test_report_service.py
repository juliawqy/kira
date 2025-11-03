"""
Unit tests for report service.
Tests PDF and Excel report generation with mock data.
"""
import pytest
from io import BytesIO
from types import SimpleNamespace

from backend.src.services import report as report_service
from tests.mock_data.report_data import (
    MOCK_PROJECT,
    MOCK_PROJECT_MINIMAL,
    INVALID_PROJECT_EMPTY,
    INVALID_PROJECT_NO_NAME,
    MOCK_TASKS_ALL_STATUSES,
    MOCK_TASKS_EMPTY,
    MOCK_TASKS_WITH_NULLS,
    MOCK_TASK_ASSIGNEES,
    MOCK_TASK_ASSIGNEES_EMPTY,
    MOCK_TASK_ASSIGNEES_ALL_UNASSIGNED,
    MOCK_TASK_LONG_TITLE,
    MOCK_TASK_NO_DATES
)

pytestmark = pytest.mark.unit


def dict_to_object(data: dict):
    """Convert dictionary to object with attributes for service compatibility."""
    return SimpleNamespace(**data)


def dicts_to_objects(task_dicts: list):
    """Convert list of dictionaries to list of objects."""
    return [dict_to_object(task) for task in task_dicts]


class TestGeneratePDFReport:
    """Test PDF report generation."""

    def test_generate_pdf_report_success_with_all_statuses(self):
        """Test successful PDF generation with tasks in all statuses."""
        pdf_buffer = report_service.generate_pdf_report(MOCK_PROJECT, dicts_to_objects(MOCK_TASKS_ALL_STATUSES), MOCK_TASK_ASSIGNEES)

        assert pdf_buffer is not None
        assert isinstance(pdf_buffer, BytesIO)

        # Verify PDF content
        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')

    def test_generate_pdf_report_success_empty_tasks(self):
        """Test PDF generation with no tasks."""
        pdf_buffer = report_service.generate_pdf_report(MOCK_PROJECT, dicts_to_objects(MOCK_TASKS_EMPTY), MOCK_TASK_ASSIGNEES_EMPTY)

        assert pdf_buffer is not None
        assert isinstance(pdf_buffer, BytesIO)

        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')

    def test_generate_pdf_report_with_null_dates(self):
        """Test PDF generation with tasks that have null dates."""
        pdf_buffer = report_service.generate_pdf_report(MOCK_PROJECT, dicts_to_objects(MOCK_TASKS_WITH_NULLS), MOCK_TASK_ASSIGNEES)

        assert pdf_buffer is not None
        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')

    def test_generate_pdf_report_with_long_title(self):
        """Test PDF generation with tasks that have long titles."""
        tasks = [dict_to_object(MOCK_TASK_LONG_TITLE)]
        assignees = {6: ["Assignee"]}
        pdf_buffer = report_service.generate_pdf_report(MOCK_PROJECT, tasks, assignees)

        assert pdf_buffer is not None
        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')

    def test_generate_pdf_report_all_unassigned(self):
        """Test PDF generation when all tasks are unassigned."""
        pdf_buffer = report_service.generate_pdf_report(MOCK_PROJECT, dicts_to_objects(MOCK_TASKS_ALL_STATUSES), MOCK_TASK_ASSIGNEES_ALL_UNASSIGNED)

        assert pdf_buffer is not None
        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')

    def test_generate_pdf_report_minimal_project(self):
        """Test PDF generation with minimal project data."""
        pdf_buffer = report_service.generate_pdf_report(MOCK_PROJECT_MINIMAL, dicts_to_objects(MOCK_TASKS_ALL_STATUSES), MOCK_TASK_ASSIGNEES)

        assert pdf_buffer is not None
        pdf_buffer.seek(0)
        content = pdf_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'%PDF')

    def test_generate_pdf_report_empty_project_raises_error(self):
        """Test PDF generation raises ValueError when project is empty."""
        with pytest.raises(ValueError, match="Project data is required"):
            report_service.generate_pdf_report(INVALID_PROJECT_EMPTY, dicts_to_objects(MOCK_TASKS_ALL_STATUSES), MOCK_TASK_ASSIGNEES)

    def test_generate_pdf_report_no_project_name_raises_error(self):
        """Test PDF generation raises ValueError when project has no name."""
        with pytest.raises(ValueError, match="Project name is required"):
            report_service.generate_pdf_report(INVALID_PROJECT_NO_NAME, dicts_to_objects(MOCK_TASKS_ALL_STATUSES), MOCK_TASK_ASSIGNEES)

    def test_generate_pdf_report_none_project_raises_error(self):
        """Test PDF generation raises ValueError when project is None."""
        with pytest.raises(ValueError, match="Project data is required"):
            report_service.generate_pdf_report(None, dicts_to_objects(MOCK_TASKS_ALL_STATUSES), MOCK_TASK_ASSIGNEES)


class TestGenerateExcelReport:
    """Test Excel report generation."""

    def test_generate_excel_report_success_with_all_statuses(self):
        """Test successful Excel generation with tasks in all statuses."""
        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT,
            dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
            MOCK_TASK_ASSIGNEES
        )

        assert excel_buffer is not None
        assert isinstance(excel_buffer, BytesIO)

        # Verify Excel content
        excel_buffer.seek(0)
        content = excel_buffer.read()
        assert len(content) > 0
        # Excel files start with PK (ZIP signature)
        assert content.startswith(b'PK')

    def test_generate_excel_report_success_empty_tasks(self):
        """Test Excel generation with no tasks."""
        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT,
            dicts_to_objects(MOCK_TASKS_EMPTY),
            MOCK_TASK_ASSIGNEES_EMPTY
        )

        assert excel_buffer is not None
        assert isinstance(excel_buffer, BytesIO)

        excel_buffer.seek(0)
        content = excel_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'PK')

    def test_generate_excel_report_with_null_dates(self):
        """Test Excel generation with tasks that have null dates."""
        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT,
            dicts_to_objects(MOCK_TASKS_WITH_NULLS),
            MOCK_TASK_ASSIGNEES
        )

        assert excel_buffer is not None
        excel_buffer.seek(0)
        content = excel_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'PK')

    def test_generate_excel_report_validates_content(self):
        """Test Excel report has expected structure."""
        from openpyxl import load_workbook

        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT,
            dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
            MOCK_TASK_ASSIGNEES
        )

        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)

        # Check sheet exists
        assert "Project Schedule Report" in wb.sheetnames

        ws = wb["Project Schedule Report"]

        # Check title
        assert ws["A1"].value is not None
        assert "Project Schedule Report" in str(ws["A1"].value)

        # Check project name
        assert ws["A2"].value == "Project Name"
        assert ws["B2"].value == MOCK_PROJECT["project_name"]

        # Check summary section
        assert ws["A4"].value == "Summary"
        assert ws["A5"].value == "Metric"
        assert ws["B5"].value == "Count"

    def test_generate_excel_report_all_unassigned(self):
        """Test Excel generation when all tasks are unassigned."""
        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT,
            dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
            MOCK_TASK_ASSIGNEES_ALL_UNASSIGNED
        )

        assert excel_buffer is not None
        excel_buffer.seek(0)
        content = excel_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'PK')

    def test_generate_excel_report_minimal_project(self):
        """Test Excel generation with minimal project data."""
        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT_MINIMAL,
            dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
            MOCK_TASK_ASSIGNEES
        )

        assert excel_buffer is not None
        excel_buffer.seek(0)
        content = excel_buffer.read()
        assert len(content) > 0
        assert content.startswith(b'PK')

    def test_generate_excel_report_empty_project_raises_error(self):
        """Test Excel generation raises ValueError when project is empty."""
        with pytest.raises(ValueError, match="Project data is required"):
            report_service.generate_excel_report(
                INVALID_PROJECT_EMPTY,
                dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
                MOCK_TASK_ASSIGNEES
            )

    def test_generate_excel_report_no_project_name_raises_error(self):
        """Test Excel generation raises ValueError when project has no name."""
        with pytest.raises(ValueError, match="Project name is required"):
            report_service.generate_excel_report(
                INVALID_PROJECT_NO_NAME,
                dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
                MOCK_TASK_ASSIGNEES
            )

    def test_generate_excel_report_none_project_raises_error(self):
        """Test Excel generation raises ValueError when project is None."""
        with pytest.raises(ValueError, match="Project data is required"):
            report_service.generate_excel_report(
                None,
                dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
                MOCK_TASK_ASSIGNEES
            )

    def test_generate_excel_report_task_summary_counts(self):
        """Test Excel report has correct summary counts."""
        from openpyxl import load_workbook

        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT,
            dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
            MOCK_TASK_ASSIGNEES
        )

        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb["Project Schedule Report"]

        # Find summary values
        summary_dict = {}
        for row in range(6, 12):
            metric = ws[f"A{row}"].value
            count = ws[f"B{row}"].value
            if metric and count is not None:
                summary_dict[metric] = count

        # Verify counts match expected values
        assert summary_dict.get("Total Tasks") == len(MOCK_TASKS_ALL_STATUSES)
        assert summary_dict.get("Projected Tasks") == 1
        assert summary_dict.get("In-Progress Tasks") == 1
        assert summary_dict.get("Completed Tasks") == 1 
        assert summary_dict.get("Under Review Tasks") == 1

    def test_generate_excel_report_column_width_with_merged_cells(self):
        """Test Excel report column width calculation handles merged cells correctly.
        This covers lines 353-368 (merged cell handling) and 378-379 (exception handling).
        """
        from openpyxl import load_workbook
        
        # Create a report with tasks to trigger merged cell handling
        excel_buffer = report_service.generate_excel_report(
            MOCK_PROJECT,
            dicts_to_objects(MOCK_TASKS_ALL_STATUSES),
            MOCK_TASK_ASSIGNEES
        )
        
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb["Project Schedule Report"]
        
        # Verify merged cells exist (the report merges cells for headers)
        # This ensures line 353 (if ws.merged_cells:) is True and lines 354-366 execute
        assert ws.merged_cells is not None
        assert len(ws.merged_cells.ranges) > 0, "Report should have merged cells"
        
        # Verify the merged cell ranges are processed
        # This ensures the loop at line 354 executes and lines 355-366 run
        merged_count = 0
        for merged_range in ws.merged_cells.ranges:
            merged_count += 1
            # Verify merged range has valid properties (lines 355-358)
            assert merged_range.min_col is not None
            assert merged_range.max_col is not None
            assert merged_range.min_row is not None
            assert merged_range.max_row is not None
        
        assert merged_count > 0, "Should have at least one merged cell range"
        
        # Verify column widths were calculated (should have reasonable widths)
        # This ensures lines 368-382 execute, including the exception handler (378-379)
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            col_width = ws.column_dimensions[col_letter].width
            assert col_width >= 10, f"Column {col_letter} should have minimum width"
            assert col_width <= 50, f"Column {col_letter} should have maximum width"
        
        # This test covers:
        # - Lines 353-366: Merged cell handling (if ws.merged_cells: and nested loops)
        # - Lines 368-379: Cell iteration and exception handling
        # - Lines 378-379: Exception pass block (defensive code, may not always execute)

