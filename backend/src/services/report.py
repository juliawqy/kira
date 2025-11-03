"""
Report generation service for project schedule reports.
Supports PDF and Excel export formats.
"""
from __future__ import annotations

from datetime import date, datetime
from typing import List, Dict, Any, Optional
from io import BytesIO

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.src.database.models.task import Task
from backend.src.enums.task_status import TaskStatus


def _get_tasks_by_status(tasks: List[Task], status: str) -> List[Task]:
    """Filter tasks by status."""
    return [task for task in tasks if task.status == status]


def _format_date(d: Optional[date]) -> str:
    """Format date for display."""
    return d.strftime("%Y-%m-%d") if d else "N/A"


def _get_task_summary_data(tasks: List[Task]) -> Dict[str, Any]:
    """Extract summary data from tasks grouped by status."""
    to_do = _get_tasks_by_status(tasks, TaskStatus.TO_DO.value)
    in_progress = _get_tasks_by_status(tasks, TaskStatus.IN_PROGRESS.value)
    completed = _get_tasks_by_status(tasks, TaskStatus.COMPLETED.value)
    blocked = _get_tasks_by_status(tasks, TaskStatus.BLOCKED.value)
    
    return {
        "projected": to_do,  # "Projected" tasks are To-do
        "completed": completed,
        "in_progress": in_progress,
        "under_review": blocked,  # "Under review" tasks are Blocked
        "total": len(tasks),
    }


def _get_assignees_string(task: Task, task_assignees: Dict[int, List[str]]) -> str:
    """Get comma-separated list of assignee names for a task."""
    assignee_names = task_assignees.get(task.id, [])
    return ", ".join(assignee_names) if assignee_names else "Unassigned"


def generate_pdf_report(
    project: Dict[str, Any],
    tasks: List[Task],
    task_assignees: Dict[int, List[str]]
) -> BytesIO:
    """
    Generate a PDF report for a project showing task schedule.
    
    Args:
        project: Dictionary containing project information with at least 'project_name' key
        tasks: List of Task objects to include in the report
        task_assignees: Dictionary mapping task_id to list of assignee names
    
    Returns:
        BytesIO buffer with PDF content.
    """
    if not project:
        raise ValueError("Project data is required")
    
    if not project.get('project_name'):
        raise ValueError("Project name is required")
    
    all_tasks = list(tasks)
    
    summary = _get_task_summary_data(all_tasks)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#283593'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    story.append(Paragraph(f"Project Schedule Report: {project['project_name']}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    summary_data = [
        ["Metric", "Count"],
        ["Total Tasks", str(summary["total"])],
        ["Projected Tasks", str(len(summary["projected"]))],
        ["In-Progress Tasks", str(len(summary["in_progress"]))],
        ["Completed Tasks", str(len(summary["completed"]))],
        ["Under Review Tasks", str(len(summary["under_review"]))],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3949ab')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    status_groups = [
        ("Projected Tasks", summary["projected"], colors.HexColor('#4caf50')),
        ("In-Progress Tasks", summary["in_progress"], colors.HexColor('#ff9800')),
        ("Completed Tasks", summary["completed"], colors.HexColor('#2196f3')),
        ("Under Review Tasks", summary["under_review"], colors.HexColor('#f44336')),
    ]
    
    for group_title, group_tasks, color in status_groups:
        if not group_tasks:
            continue
            
        story.append(Paragraph(group_title, heading_style))
        
        task_data = [["ID", "Title", "Priority", "Start Date", "Deadline", "Assignees"]]
        
        for task in group_tasks:
            assignees = _get_assignees_string(task, task_assignees)
            task_data.append([
                str(task.id),
                task.title[:40] + "..." if task.title and len(task.title) > 40 else (task.title or "N/A"),
                str(task.priority),
                _format_date(task.start_date),
                _format_date(task.deadline),
                assignees[:30] + "..." if len(assignees) > 30 else assignees,
            ])
        
        task_table = Table(task_data, colWidths=[0.5*inch, 2.5*inch, 0.7*inch, 1*inch, 1*inch, 1.8*inch])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(task_table)
        story.append(Spacer(1, 0.2*inch))
    
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(
        f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        styles['Normal']
    ))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_excel_report(
    project: Dict[str, Any],
    tasks: List[Task],
    task_assignees: Dict[int, List[str]]
) -> BytesIO:
    """
    Generate an Excel report for a project showing task schedule.
    All data is in a single sheet with tasks grouped by status.
    
    Args:
        project: Dictionary containing project information with at least 'project_name' key
        tasks: List of Task objects to include in the report
        task_assignees: Dictionary mapping task_id to list of assignee names
    
    Returns:
        BytesIO buffer with Excel content.
    """
    if not project:
        raise ValueError("Project data is required")
    
    if not project.get('project_name'):
        raise ValueError("Project name is required")
    

    all_tasks = list(tasks)
    
    summary = _get_task_summary_data(all_tasks)
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Schedule Report"
        # Header style
    header_fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws['A1'] = "Project Schedule Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = "Project Name"
    ws['A2'].font = Font(bold=True)
    ws['B2'] = project['project_name']
    ws['B2'].font = Font(bold=True)
    ws.merge_cells('B2:H2')
    
    ws['A4'] = "Summary"
    ws['A4'].font = Font(bold=True, size=14)
    ws.merge_cells('A4:H4')
    
    ws['A5'] = "Metric"
    ws['B5'] = "Count"
    ws['A5'].fill = header_fill
    ws['B5'].fill = header_fill
    ws['A5'].font = header_font
    ws['B5'].font = header_font
    ws['A5'].border = border
    ws['B5'].border = border
    
    summary_rows = [
        ["Total Tasks", summary["total"]],
        ["Projected Tasks", len(summary["projected"])],
        ["In-Progress Tasks", len(summary["in_progress"])],
        ["Completed Tasks", len(summary["completed"])],
        ["Under Review Tasks", len(summary["under_review"])],
    ]
    
    for idx, (metric, count) in enumerate(summary_rows, start=6):
        ws[f'A{idx}'] = metric
        ws[f'B{idx}'] = count
        ws[f'A{idx}'].border = border
        ws[f'B{idx}'].border = border
        if idx % 2 == 0:
            ws[f'A{idx}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            ws[f'B{idx}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    

    current_row = 12
    
    status_groups = [
        ("PROJECTED TASKS", summary["projected"], "4CAF50"),
        ("IN-PROGRESS TASKS", summary["in_progress"], "FF9800"),
        ("COMPLETED TASKS", summary["completed"], "2196F3"),
        ("UNDER REVIEW TASKS", summary["under_review"], "F44336"),
    ]
    
    headers = ["ID", "Title", "Description", "Priority", "Start Date", "Deadline", "Assignees", "Tag"]
    
    for status_title, task_list, color_hex in status_groups:
        if not task_list:
            continue
        
        status_cell = ws.cell(row=current_row, column=1, value=status_title)
        status_cell.font = Font(bold=True, size=13)
        status_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        status_cell.font = Font(bold=True, color="FFFFFF", size=13)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        status_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        for task in task_list:
            assignees = _get_assignees_string(task, task_assignees)
            row_data = [
                task.id,
                task.title or "N/A",
                (task.description or "N/A")[:100],
                task.priority,
                _format_date(task.start_date),
                _format_date(task.deadline),
                assignees,
                task.tag or "N/A",
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                if current_row % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            current_row += 1

        current_row += 1
    
    from openpyxl.utils import get_column_letter
    
    for col_idx in range(1, 9):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, current_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

